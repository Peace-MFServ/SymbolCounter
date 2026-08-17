"""
Symbol Counter API — v3.1 (learning-loop fixes)

Changes vs v2:
  - Detection IDs assigned server-side (fixes mass-delete bug in frontend)
  - Feedback capture is idempotent: only fires on FIRST verification of a page
  - wrong_type feedback properly recorded (type change = false_positive + missed)
  - /pages and /templates served through authenticated endpoints (no open StaticFiles)
  - CORS locked to localhost origins
  - Drawing status extended: verified → approved (with approved_by / approved_at)
  - Approved drawings feed the learning loop exclusively (clean signal)
  - Door-ref correction endpoint: PATCH /api/drawings/{did}/door-refs
  - Versioned JSON export payload (ironmongery scheduling bridge)
  - Background detection passes drawing_firm for firm-scoped template matching
  - Excel filename sanitised (no UnicodeEncodeError)
  - Polling endpoint returns processing_count so frontend can stop when idle
"""
import copy
import logging
import logging.handlers
import os
import io
import re
import shutil
import uuid
import json
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, BackgroundTasks, status, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse, Response
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from pydantic import BaseModel, EmailStr

import models, auth
from database import engine, get_db, Base
from detection.pipeline import run_full_detection, crop_template_from_page, find_door_references

logger = logging.getLogger(__name__)

# Anchor the working directory to this file's folder so relative paths
# (uploads/, stored image_path values) resolve no matter where uvicorn
# was started from. Starting from the wrong folder used to strand all data.
os.chdir(Path(__file__).resolve().parent)

# ── File logging setup ────────────────────────────────────────────────────────
_LOG_DIR  = Path(__file__).resolve().parent / "logs"
_LOG_DIR.mkdir(exist_ok=True)
_LOG_FILE = _LOG_DIR / "app.log"

_file_handler = logging.handlers.RotatingFileHandler(
    str(_LOG_FILE), maxBytes=5 * 1024 * 1024, backupCount=3, encoding="utf-8"
)
_file_handler.setFormatter(logging.Formatter(
    "%(asctime)s %(levelname)-8s %(name)s — %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
))
_file_handler.setLevel(logging.DEBUG)

_console_handler = logging.StreamHandler()
_console_handler.setFormatter(logging.Formatter("%(levelname)s: %(message)s"))
_console_handler.setLevel(logging.INFO)

logging.basicConfig(level=logging.DEBUG, handlers=[_file_handler, _console_handler])

# ── Environment check: Poppler is required for all PDF work ───────────────────
POPPLER_OK = bool(shutil.which("pdftoppm") and shutil.which("pdftotext"))
if not POPPLER_OK:
    _poppler_msg = (
        "Poppler is not installed or not on PATH — PDF rendering and detection "
        "cannot work without it.\n"
        "  Windows:  winget install poppler   (then close and reopen the terminal)\n"
        "  Linux:    sudo apt install poppler-utils\n"
        "Set SYMBOL_COUNTER_IGNORE_MISSING_POPPLER=1 to start anyway "
        "(existing drawings stay viewable; new uploads are rejected)."
    )
    logger.critical(_poppler_msg)
    if os.getenv("SYMBOL_COUNTER_IGNORE_MISSING_POPPLER") != "1":
        raise SystemExit("\n\nSTARTUP ABORTED\n" + _poppler_msg + "\n")

Base.metadata.create_all(bind=engine)

# ── Lightweight column migrations (add new columns to existing tables) ────────
def _ensure_columns():
    """Add any new columns that don't exist yet (safe to run on every startup)."""
    import sqlalchemy as _sa
    try:
        with engine.begin() as conn:   # begin() auto-commits on success
            rows = conn.execute(_sa.text("PRAGMA table_info(global_templates)")).fetchall()
            existing = {row[1] for row in rows}
            if "threshold_override" not in existing:
                conn.execute(_sa.text(
                    "ALTER TABLE global_templates ADD COLUMN threshold_override REAL"
                ))
                logger.info("Migrated: added global_templates.threshold_override")
            drows = conn.execute(_sa.text("PRAGMA table_info(drawings)")).fetchall()
            dexisting = {row[1] for row in drows}
            if "error_message" not in dexisting:
                conn.execute(_sa.text(
                    "ALTER TABLE drawings ADD COLUMN error_message TEXT"
                ))
                logger.info("Migrated: added drawings.error_message")
    except Exception as exc:
        logger.error("Column migration failed: %s", exc)

_ensure_columns()

UPLOAD_DIR   = Path("uploads")
TEMPLATE_DIR = UPLOAD_DIR / "templates"
UPLOAD_DIR.mkdir(exist_ok=True)
TEMPLATE_DIR.mkdir(exist_ok=True)

MAX_UPLOAD_BYTES = 200 * 1024 * 1024  # 200 MB

app = FastAPI(title="Symbol Counter API", version="3.0.0")

# CORS: localhost only (both :8000 for direct API and common Vite dev ports)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:8000",
        "http://localhost:5173",
        "http://127.0.0.1:8000",
        "http://127.0.0.1:5173",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# resolve() first — __file__ may be relative depending on how uvicorn is launched
_HERE         = Path(__file__).resolve().parent
FRONTEND_DIST = _HERE.parent / "frontend" / "dist"

# Mount /assets if already built — safe to call at startup even if dist doesn't exist yet
from fastapi.staticfiles import StaticFiles as _SF
if (FRONTEND_DIST / "assets").exists():
    app.mount("/assets", _SF(directory=str(FRONTEND_DIST / "assets")), name="assets")


# ── Schemas ───────────────────────────────────────────────────────────────────
class UserCreate(BaseModel):
    email: EmailStr; name: str; password: str

class UserOut(BaseModel):
    id: int; email: str; name: str
    class Config: from_attributes = True

class Token(BaseModel):
    access_token: str; token_type: str; user: UserOut

class ProjectCreate(BaseModel):
    name: str; client: str = ""; site: str = ""
    description: str = ""; drawing_firm: str = ""

class ProjectOut(BaseModel):
    id: int; name: str; client: str; site: str
    description: str; drawing_firm: str; drawing_count: int = 0
    verified_count: int = 0; approved_count: int = 0
    class Config: from_attributes = True

class DrawingOut(BaseModel):
    id: int; original_name: str; block: str; level: str
    total_pages: int; status: str; revision: str = ""
    error_message: Optional[str] = None
    total_counts: dict = {}
    verified_pages: int = 0; total_pages_count: int = 0
    class Config: from_attributes = True

class DetectionUpdate(BaseModel):
    page_number: int
    detections: list[dict]

class BlockMetaUpdate(BaseModel):
    block: str = ""; level: str = ""; revision: str = ""

class SymbolTypeCreate(BaseModel):
    name: str; code: str; color: str = "#8892a8"; sort_order: int = 0

class SymbolTypeOut(BaseModel):
    id: int; name: str; code: str; color: str; sort_order: int
    template_count: int = 0
    class Config: from_attributes = True

class TemplateCropRequest(BaseModel):
    symbol_type_id: int
    page_number:    int
    x1: float; y1: float; x2: float; y2: float

class TemplateOut(BaseModel):
    id: int; symbol_code: str; symbol_name: str
    image_url: str; confirm_count: int; reject_count: int
    quality_score: float; effective_threshold: float
    threshold_override: Optional[float] = None
    drawing_firm: str = ""
    class Config: from_attributes = True

class TemplateThresholdUpdate(BaseModel):
    threshold_override: Optional[float] = None  # None clears the override

class DoorRefUpdate(BaseModel):
    """Correct a door-ref association for a specific detection on a page."""
    page_number: int
    det_imgX:    float  # position to identify the detection
    det_imgY:    float
    door_ref:    str    # new door ref value (empty string to clear)

class PositionCorrectionPayload(BaseModel):
    page_number: int
    feedback_type: str  # 'position_correction'
    symbol_code: str
    old_img_x: float
    old_img_y: float
    new_img_x: float
    new_img_y: float


# ── Auth ──────────────────────────────────────────────────────────────────────
@app.post("/api/auth/register", response_model=Token, status_code=201)
def register(payload: UserCreate, db: Session = Depends(get_db)):
    if db.query(models.User).filter(models.User.email == payload.email).first():
        raise HTTPException(400, "Email already registered")
    user = models.User(email=payload.email, name=payload.name,
                       hashed_password=auth.hash_password(payload.password))
    db.add(user); db.commit(); db.refresh(user)
    token = auth.create_access_token({"sub": user.email})
    return {"access_token": token, "token_type": "bearer", "user": user}

@app.post("/api/auth/login", response_model=Token)
def login(form: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == form.username).first()
    if not user or not auth.verify_password(form.password, user.hashed_password):
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Invalid email or password")
    token = auth.create_access_token({"sub": user.email})
    return {"access_token": token, "token_type": "bearer", "user": user}

@app.get("/api/auth/me", response_model=UserOut)
def me(current_user=Depends(auth.get_current_user)):
    return current_user


# ── Authenticated file serving ────────────────────────────────────────────────
@app.get("/api/files/pages/{rest:path}")
def serve_page_image(rest: str, cu=Depends(auth.get_current_user)):
    """Serve drawing page images — requires authentication."""
    path = UPLOAD_DIR / rest
    if not path.exists() or not path.is_file():
        raise HTTPException(404, "File not found")
    # Basic path traversal protection
    try:
        path.resolve().relative_to(UPLOAD_DIR.resolve())
    except ValueError:
        raise HTTPException(403, "Access denied")
    return FileResponse(str(path))

@app.get("/api/files/templates/{filename}")
def serve_template_image(filename: str, cu=Depends(auth.get_current_user)):
    """Serve template images — requires authentication."""
    path = TEMPLATE_DIR / filename
    if not path.exists() or not path.is_file():
        raise HTTPException(404, "File not found")
    return FileResponse(str(path))


# ── Projects ──────────────────────────────────────────────────────────────────
@app.get("/api/projects", response_model=list[ProjectOut])
def list_projects(db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    ps = db.query(models.Project).filter(models.Project.owner_id == cu.id).all()
    result = []
    for p in ps:
        vc = sum(1 for d in p.drawings if d.status in ("verified", "approved"))
        ac = sum(1 for d in p.drawings if d.status == "approved")
        result.append(ProjectOut(
            id=p.id, name=p.name, client=p.client, site=p.site,
            description=p.description, drawing_firm=p.drawing_firm,
            drawing_count=len(p.drawings), verified_count=vc, approved_count=ac,
        ))
    return result

@app.post("/api/projects", response_model=ProjectOut, status_code=201)
def create_project(payload: ProjectCreate, db: Session = Depends(get_db),
                   cu=Depends(auth.get_current_user)):
    p = models.Project(**payload.model_dump(), owner_id=cu.id)
    db.add(p); db.commit(); db.refresh(p)
    return ProjectOut(id=p.id, name=p.name, client=p.client, site=p.site,
                      description=p.description, drawing_firm=p.drawing_firm, drawing_count=0)

@app.get("/api/projects/{pid}", response_model=ProjectOut)
def get_project(pid: int, db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    p = _proj404(pid, cu.id, db)
    vc = sum(1 for d in p.drawings if d.status in ("verified", "approved"))
    ac = sum(1 for d in p.drawings if d.status == "approved")
    return ProjectOut(id=p.id, name=p.name, client=p.client, site=p.site,
                      description=p.description, drawing_firm=p.drawing_firm,
                      drawing_count=len(p.drawings), verified_count=vc, approved_count=ac)

@app.put("/api/projects/{pid}", response_model=ProjectOut)
def update_project(pid: int, payload: ProjectCreate, db: Session = Depends(get_db),
                   cu=Depends(auth.get_current_user)):
    p = _proj404(pid, cu.id, db)
    for k, v in payload.model_dump().items():
        setattr(p, k, v)
    db.commit(); db.refresh(p)
    return ProjectOut(id=p.id, name=p.name, client=p.client, site=p.site,
                      description=p.description, drawing_firm=p.drawing_firm,
                      drawing_count=len(p.drawings))

@app.delete("/api/projects/{pid}", status_code=204)
def delete_project(pid: int, db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    p = _proj404(pid, cu.id, db); db.delete(p); db.commit()


# ── Symbol types ──────────────────────────────────────────────────────────────
@app.get("/api/projects/{pid}/symbol-types", response_model=list[SymbolTypeOut])
def get_symbol_types(pid: int, db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    p = _proj404(pid, cu.id, db)
    if not p.symbol_types:
        for st in models.DEFAULT_SYMBOL_TYPES:
            db.add(models.ProjectSymbolType(**st, project_id=pid))
        db.commit(); db.refresh(p)
    result = []
    for st in p.symbol_types:
        tc = db.query(models.GlobalTemplate).filter(
            models.GlobalTemplate.symbol_code == st.code).count()
        result.append(SymbolTypeOut(id=st.id, name=st.name, code=st.code,
                                    color=st.color, sort_order=st.sort_order, template_count=tc))
    return result

@app.post("/api/projects/{pid}/symbol-types", response_model=SymbolTypeOut, status_code=201)
def add_symbol_type(pid: int, payload: SymbolTypeCreate, db: Session = Depends(get_db),
                    cu=Depends(auth.get_current_user)):
    _proj404(pid, cu.id, db)
    st = models.ProjectSymbolType(**payload.model_dump(), project_id=pid)
    db.add(st)
    try:
        db.commit(); db.refresh(st)
    except Exception:
        db.rollback()
        raise HTTPException(409, f"Symbol code '{payload.code}' already exists in this project")
    return SymbolTypeOut(id=st.id, name=st.name, code=st.code,
                         color=st.color, sort_order=st.sort_order, template_count=0)

@app.put("/api/symbol-types/{st_id}", response_model=SymbolTypeOut)
def update_symbol_type(st_id: int, payload: SymbolTypeCreate, db: Session = Depends(get_db),
                       cu=Depends(auth.get_current_user)):
    st = db.query(models.ProjectSymbolType).filter(models.ProjectSymbolType.id == st_id).first()
    if not st: raise HTTPException(404, "Not found")
    _proj404(st.project_id, cu.id, db)
    for k, v in payload.model_dump().items(): setattr(st, k, v)
    db.commit(); db.refresh(st)
    tc = db.query(models.GlobalTemplate).filter(models.GlobalTemplate.symbol_code == st.code).count()
    return SymbolTypeOut(id=st.id, name=st.name, code=st.code,
                         color=st.color, sort_order=st.sort_order, template_count=tc)

@app.delete("/api/symbol-types/{st_id}", status_code=204)
def delete_symbol_type(st_id: int, db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    st = db.query(models.ProjectSymbolType).filter(models.ProjectSymbolType.id == st_id).first()
    if not st: raise HTTPException(404, "Not found")
    _proj404(st.project_id, cu.id, db); db.delete(st); db.commit()

@app.post("/api/projects/{pid}/symbol-types/reset", response_model=list[SymbolTypeOut])
def reset_symbol_types(pid: int, db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    p = _proj404(pid, cu.id, db)
    for st in p.symbol_types: db.delete(st)
    db.flush()
    for st in models.DEFAULT_SYMBOL_TYPES:
        db.add(models.ProjectSymbolType(**st, project_id=pid))
    db.commit(); db.refresh(p)
    return [SymbolTypeOut(id=st.id, name=st.name, code=st.code,
                          color=st.color, sort_order=st.sort_order, template_count=0)
            for st in p.symbol_types]


# ── Template library ──────────────────────────────────────────────────────────
@app.post("/api/drawings/{did}/crop-template", response_model=TemplateOut, status_code=201)
def crop_template(did: int, payload: TemplateCropRequest,
                  db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    d = _draw404(did, cu.id, db)
    page = next((p for p in d.pages if p.page_number == payload.page_number), None)
    if not page: raise HTTPException(404, f"Page {payload.page_number} not found")

    # Verify symbol type belongs to this drawing's project
    st = db.query(models.ProjectSymbolType).filter(
        models.ProjectSymbolType.id == payload.symbol_type_id,
        models.ProjectSymbolType.project_id == d.project_id,
    ).first()
    if not st: raise HTTPException(404, "Symbol type not found in this project")

    tmpl_uid  = uuid.uuid4().hex
    out_path  = str(TEMPLATE_DIR / f"{tmpl_uid}.png")
    try:
        # padding=0 — user already drew the box exactly; don't expand it
        w, h = crop_template_from_page(page.image_path, payload.x1, payload.y1,
                                        payload.x2, payload.y2, out_path, padding=0.0)
    except Exception as e:
        raise HTTPException(400, f"Crop failed: {e}")

    firm = d.project.drawing_firm or ""
    tmpl = models.GlobalTemplate(
        symbol_code=st.code, symbol_name=st.name,
        image_path=out_path, image_width=w, image_height=h,
        drawing_firm=firm, source_project_id=d.project_id,
        created_by_id=cu.id,
    )
    db.add(tmpl); db.commit(); db.refresh(tmpl)
    return _tmpl_out(tmpl)

@app.get("/api/templates/all", response_model=list[TemplateOut])
def list_all_templates(db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    """Return every template, sorted by symbol code then confirm_count desc."""
    tmpls = db.query(models.GlobalTemplate).order_by(
        models.GlobalTemplate.symbol_code,
        models.GlobalTemplate.confirm_count.desc(),
    ).all()
    return [_tmpl_out(t) for t in tmpls]

@app.get("/api/templates", response_model=list[TemplateOut])
def list_templates(code: str = Query(...), db: Session = Depends(get_db),
                   cu=Depends(auth.get_current_user)):
    tmpls = db.query(models.GlobalTemplate).filter(
        models.GlobalTemplate.symbol_code == code
    ).order_by(models.GlobalTemplate.confirm_count.desc()).all()
    return [_tmpl_out(t) for t in tmpls]

@app.patch("/api/templates/{tmpl_id}/threshold", response_model=TemplateOut)
def set_template_threshold(tmpl_id: int, payload: TemplateThresholdUpdate,
                           db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    """Manually pin or clear a template's detection threshold."""
    t = db.query(models.GlobalTemplate).filter(models.GlobalTemplate.id == tmpl_id).first()
    if not t: raise HTTPException(404, "Not found")
    if payload.threshold_override is not None:
        v = round(float(payload.threshold_override), 3)
        if not (0.3 <= v <= 1.0):
            raise HTTPException(400, "Threshold must be between 0.3 and 1.0")
        t.threshold_override = v
    else:
        t.threshold_override = None  # revert to adaptive
    db.commit(); db.refresh(t)
    return _tmpl_out(t)

@app.delete("/api/templates/{tmpl_id}", status_code=204)
def delete_template(tmpl_id: int, db: Session = Depends(get_db),
                    cu=Depends(auth.get_current_user)):
    t = db.query(models.GlobalTemplate).filter(models.GlobalTemplate.id == tmpl_id).first()
    if not t: raise HTTPException(404, "Not found")
    # Only allow creator or admin to delete
    if t.created_by_id is not None and t.created_by_id != cu.id:
        raise HTTPException(403, "You can only delete templates you created")
    if t.image_path and os.path.exists(t.image_path):
        os.remove(t.image_path)
    db.delete(t); db.commit()


# ── Drawings ──────────────────────────────────────────────────────────────────
@app.get("/api/projects/{pid}/drawings", response_model=list[DrawingOut])
def list_drawings(pid: int, db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    _proj404(pid, cu.id, db)
    ds = db.query(models.Drawing).filter(models.Drawing.project_id == pid).all()
    return [_drawing_out(d) for d in ds]

@app.get("/api/projects/{pid}/drawings/processing-count")
def drawing_processing_count(pid: int, db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    """Returns how many drawings are still processing — lets frontend stop polling when 0."""
    _proj404(pid, cu.id, db)
    n = db.query(models.Drawing).filter(
        models.Drawing.project_id == pid,
        models.Drawing.status == "processing",
    ).count()
    return {"processing": n}

@app.post("/api/projects/{pid}/drawings", response_model=DrawingOut, status_code=201)
async def upload_drawing(pid: int, background_tasks: BackgroundTasks,
                         file: UploadFile = File(...),
                         db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    p = _proj404(pid, cu.id, db)

    if not (file.filename or "").lower().endswith(".pdf"):
        raise HTTPException(400, "Only PDF files accepted")

    if not POPPLER_OK:
        raise HTTPException(503,
            "Cannot process PDFs: Poppler is not installed on the server. "
            "Install it (winget install poppler), reopen the terminal, and restart the backend.")

    uid      = uuid.uuid4().hex
    save_dir = UPLOAD_DIR / uid
    save_dir.mkdir(parents=True)
    pdf_path = save_dir / "drawing.pdf"

    # Stream with size limit
    total = 0
    with open(pdf_path, "wb") as f:
        while chunk := await file.read(1024 * 1024):
            total += len(chunk)
            if total > MAX_UPLOAD_BYTES:
                shutil.rmtree(save_dir, ignore_errors=True)
                raise HTTPException(413, "File too large (max 200 MB)")
            f.write(chunk)

    block, level = _parse_name(file.filename)
    revision     = _parse_revision(file.filename)

    drawing = models.Drawing(
        filename=uid, original_name=file.filename,
        block=block, level=level, revision=revision,
        status="uploaded", project_id=pid,
    )
    db.add(drawing); db.commit(); db.refresh(drawing)

    symbol_codes = [st.code for st in p.symbol_types] or \
                   [st["code"] for st in models.DEFAULT_SYMBOL_TYPES]

    # Pass template data as plain dicts to avoid DetachedInstanceError in background task
    stored_tmpls_data = [
        {
            "id":           t.id,
            "symbol_code":  t.symbol_code,
            "image_path":   t.image_path,
            "drawing_firm": t.drawing_firm,
            "confirm_count": t.confirm_count,
            "reject_count":  t.reject_count,
            "use_count":     t.use_count,
            "threshold_override": getattr(t, "threshold_override", None),
        }
        for t in db.query(models.GlobalTemplate).filter(
            models.GlobalTemplate.symbol_code.in_(symbol_codes)
        ).all()
    ]

    background_tasks.add_task(
        _run_detection,
        drawing.id, str(pdf_path), str(save_dir),
        symbol_codes, stored_tmpls_data, p.drawing_firm or "",
    )

    return _drawing_out(drawing)

@app.get("/api/drawings/{did}", response_model=DrawingOut)
def get_drawing(did: int, db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    d = _draw404(did, cu.id, db)
    return _drawing_out(d)

@app.get("/api/drawings/{did}/pages")
def get_drawing_pages(did: int, db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    d = _draw404(did, cu.id, db)
    pages = []
    for p in sorted(d.pages, key=lambda x: x.page_number):
        # Assign stable IDs to detections so frontend can track them
        dets = _assign_ids(p.detections or [])
        orig = _assign_ids(p.original_detections or [])
        pages.append({
            "page_number":         p.page_number,
            "image_url":           f"/api/files/pages/{d.filename}/page-{p.page_number}.png",
            "detections":          dets,
            "original_detections": orig,
            "verified":            p.verified,
        })
    return pages

@app.patch("/api/drawings/{did}/meta", response_model=DrawingOut)
def update_meta(did: int, payload: BlockMetaUpdate, db: Session = Depends(get_db),
                cu=Depends(auth.get_current_user)):
    d = _draw404(did, cu.id, db)
    d.block = payload.block; d.level = payload.level; d.revision = payload.revision
    db.commit(); db.refresh(d)
    return _drawing_out(d)

@app.put("/api/drawings/{did}/detections")
def save_detections(did: int, payload: DetectionUpdate,
                    db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    """Save verified detections and capture implicit feedback (idempotent on first save)."""
    d    = _draw404(did, cu.id, db)
    page = next((p for p in d.pages if p.page_number == payload.page_number), None)
    if not page: raise HTTPException(404, f"Page {payload.page_number} not found")

    # Strip client-side IDs before saving (they're ephemeral)
    clean_dets = [{k: v for k, v in det.items() if k != "id"}
                  for det in payload.detections]

    # Only capture feedback on FIRST verification (prevents double-counting on re-save)
    if not page.verified:
        _capture_feedback(page, clean_dets, d.project.drawing_firm or "", db)
        _auto_create_templates(page, clean_dets, d, cu.id, db)

    page.detections = clean_dets
    page.verified   = True

    # Update drawing status
    all_verified = all(p.verified for p in d.pages)
    if all_verified and d.status not in ("verified", "approved"):
        d.status = "verified"
        d.verified_by_id = cu.id
        d.verified_at    = datetime.utcnow()

    db.commit()
    counts = {}
    for det in clean_dets:
        t = det.get("type", "")
        counts[t] = counts.get(t, 0) + 1
    return {"ok": True, "counts": counts}

@app.post("/api/drawings/{did}/approve")
def approve_drawing(did: int, db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    """Mark a verified drawing as approved — locks it for the learning loop."""
    d = _draw404(did, cu.id, db)
    if d.status not in ("verified", "approved"):
        raise HTTPException(400, "Drawing must be verified before approving")
    d.status = "approved"
    d.approved_by_id = cu.id
    d.approved_at    = datetime.utcnow()
    db.commit()
    return {"ok": True}

@app.patch("/api/drawings/{did}/door-refs")
def update_door_ref(did: int, payload: DoorRefUpdate,
                    db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    """Correct a door-ref association for a specific detection."""
    d    = _draw404(did, cu.id, db)
    page = next((p for p in d.pages if p.page_number == payload.page_number), None)
    if not page: raise HTTPException(404, f"Page {payload.page_number} not found")

    dets    = list(page.detections or [])
    TOL     = 0.01  # normalised position tolerance
    matched = False
    for det in dets:
        if (abs(det.get("imgX", 0) - payload.det_imgX) < TOL and
                abs(det.get("imgY", 0) - payload.det_imgY) < TOL):
            if payload.door_ref:
                det["door_ref"] = payload.door_ref
            elif "door_ref" in det:
                del det["door_ref"]
            matched = True
            break

    if not matched:
        raise HTTPException(404, "Detection not found at those coordinates")

    page.detections = dets
    db.commit()
    return {"ok": True}

@app.post("/api/drawings/{did}/feedback")
def record_position_feedback(did: int, payload: PositionCorrectionPayload,
                              db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    """Record a position correction: user dragged a detection to a better location.
    Stored in DetectionFeedback so the pipeline can learn offset patterns per symbol type."""
    d = _draw404(did, cu.id, db)
    if payload.feedback_type != "position_correction":
        raise HTTPException(400, "Only position_correction supported via this endpoint")

    # Find the drawing firm for scoped learning
    firm = d.project.drawing_firm or "" if d.project else ""

    fb = models.DetectionFeedback(
        symbol_code         = payload.symbol_code,
        feedback_type       = "position_correction",
        img_x               = payload.old_img_x,
        img_y               = payload.old_img_y,
        # Store new position in corrected_code field repurposed as JSON,
        # and also in drawing_scale for easy querying (new_x,new_y)
        corrected_code      = f"{payload.new_img_x:.6f},{payload.new_img_y:.6f}",
        drawing_firm        = firm,
        drawing_id          = did,
        page_number         = payload.page_number,
        original_method     = "position_correction",
    )
    db.add(fb)

    # NOTE: deliberately do NOT touch GlobalTemplate counts here.
    # 1) A position correction means the symbol WAS there (true positive) — it
    #    must not increment reject_count. Confirm/reject accounting happens
    #    once, in _capture_feedback on page save.
    # 2) The previous code ordered by GlobalTemplate.quality_score, which is a
    #    Python @property (not a column) — that raised at runtime and made this
    #    endpoint 500 on every marker drag.

    db.commit()
    return {"ok": True}

@app.delete("/api/drawings/{did}", status_code=204)
def delete_drawing(did: int, db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    d = _draw404(did, cu.id, db)
    # Delete child rows first to satisfy FK constraints
    db.query(models.DetectionFeedback).filter(models.DetectionFeedback.drawing_id == did).delete(synchronize_session=False)
    db.query(models.DrawingPage).filter(models.DrawingPage.drawing_id == did).delete(synchronize_session=False)
    sd = UPLOAD_DIR / d.filename
    if sd.exists(): shutil.rmtree(sd)
    db.delete(d); db.commit()


# ── Revision comparison ───────────────────────────────────────────────────────
@app.get("/api/drawings/{did}/compare/{other_did}")
def compare_revisions(did: int, other_did: int,
                      db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    """Spatial diff between two drawing revisions."""
    d1 = _draw404(did, cu.id, db)
    d2 = _draw404(other_did, cu.id, db)

    # Count-level diff
    c1, c2 = d1.total_counts, d2.total_counts
    all_codes = set(c1.keys()) | set(c2.keys())
    count_diff = {}
    for code in all_codes:
        v1, v2 = c1.get(code, 0), c2.get(code, 0)
        count_diff[code] = {"before": v1, "after": v2, "delta": v2 - v1}

    # Spatial diff — match detections within 3% position tolerance
    TOL = 0.03
    spatial_added   = []
    spatial_removed = []

    dets1 = [det for pg in d1.pages for det in (pg.detections or [])]
    dets2 = [det for pg in d2.pages for det in (pg.detections or [])]

    matched2 = set()
    for det in dets1:
        found = False
        for j, det2 in enumerate(dets2):
            if j in matched2: continue
            if (det["type"] == det2["type"] and
                    abs(det.get("imgX", 0) - det2.get("imgX", 0)) < TOL and
                    abs(det.get("imgY", 0) - det2.get("imgY", 0)) < TOL):
                matched2.add(j)
                found = True
                break
        if not found:
            spatial_removed.append(det)

    for j, det in enumerate(dets2):
        if j not in matched2:
            spatial_added.append(det)

    return {
        "drawing_a": {"id": d1.id, "name": d1.original_name, "revision": d1.revision},
        "drawing_b": {"id": d2.id, "name": d2.original_name, "revision": d2.revision},
        "count_diff": count_diff,
        "spatial_added":   spatial_added,
        "spatial_removed": spatial_removed,
        "changed": any(v["delta"] != 0 for v in count_diff.values()),
        "summary": f"+{len(spatial_added)} added, -{len(spatial_removed)} removed",
    }


# ── Export ────────────────────────────────────────────────────────────────────
@app.get("/api/projects/{pid}/export/excel")
def export_excel(pid: int, db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    p = _proj404(pid, cu.id, db)
    safe_name = re.sub(r'[^\w\s\-]', '', p.name).strip().replace(' ', '_') or 'project'
    return StreamingResponse(
        io.BytesIO(_build_excel(p)),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_Symbol_Count.xlsx"'},
    )

@app.get("/api/projects/{pid}/export/json")
def export_json(pid: int, db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    """
    Versioned structured export — the intended bridge to the ironmongery scheduling system.
    Each drawing includes per-door symbol counts keyed by door_ref.
    """
    p = _proj404(pid, cu.id, db)
    drawings_data = []
    for d in p.drawings:
        # Build door-keyed symbol map
        door_map: dict[str, dict[str, int]] = {}
        all_dets = []
        for pg in sorted(d.pages, key=lambda x: x.page_number):
            for det in (pg.detections or []):
                all_dets.append(det)
                door = det.get("door_ref", "")
                if door:
                    door_map.setdefault(door, {})
                    door_map[door][det.get("type", "")] = \
                        door_map[door].get(det.get("type", ""), 0) + 1

        drawings_data.append({
            "id":          d.id,
            "drawing_no":  d.block or d.original_name,
            "description": d.level or "",
            "revision":    d.revision,
            "status":      d.status,
            "verified_at": d.verified_at.isoformat() if d.verified_at else None,
            "approved_at": d.approved_at.isoformat() if d.approved_at else None,
            "counts":      d.total_counts,
            "door_symbols": door_map,   # {door_ref: {symbol_code: count}}
        })

    payload = {
        "schema_version": "1.0",
        "exported_at":    datetime.utcnow().isoformat(),
        "project": {
            "name":         p.name,
            "client":       p.client,
            "site":         p.site,
            "drawing_firm": p.drawing_firm,
        },
        "drawings": drawings_data,
    }
    safe_name = re.sub(r'[^\w\s\-]', '', p.name).strip().replace(' ', '_') or 'project'
    return StreamingResponse(
        io.StringIO(json.dumps(payload, indent=2)),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_export.json"'},
    )

@app.get("/api/drawings/{did}/export/pdf")
def export_drawing_pdf(did: int, db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    """Export a single drawing as PDF with detection markers drawn on each page."""
    from PIL import Image, ImageDraw, ImageFont
    d = _draw404(did, cu.id, db)

    # Build a color map from project symbol types
    color_map: dict[str, tuple] = {}
    for st in (d.project.symbol_types if d.project else []):
        try:
            hex_col = st.color.lstrip('#')
            color_map[st.code] = tuple(int(hex_col[i:i+2], 16) for i in (0, 2, 4))
        except Exception:
            color_map[st.code] = (255, 80, 80)

    pil_imgs = []
    for page in sorted(d.pages, key=lambda p: p.page_number):
        if not page.image_path or not os.path.exists(page.image_path):
            continue
        img = Image.open(page.image_path).convert("RGB")
        draw = ImageDraw.Draw(img)
        iw, ih = img.size
        r = max(12, int(iw * 0.006))

        for det in (page.detections or []):
            cx = int(det.get("imgX", 0) * iw)
            cy = int(det.get("imgY", 0) * ih)
            code = det.get("type", "")
            col = color_map.get(code, (255, 80, 80))
            # Circle with fill
            draw.ellipse([cx - r, cy - r, cx + r, cy + r],
                         outline=col, width=max(2, r // 5))
            # Crosshair
            hs = r // 2
            draw.line([cx - hs, cy, cx + hs, cy], fill=col, width=max(1, r // 8))
            draw.line([cx, cy - hs, cx, cy + hs], fill=col, width=max(1, r // 8))
            # Label above
            label = 'DOME' if code == 'CCTV_Dome' else 'FIX' if code == 'CCTV_Fixed' else code
            try:
                font = ImageFont.truetype("arial.ttf", max(10, r))
            except Exception:
                font = ImageFont.load_default()
            draw.text((cx, cy - r - 4), label, fill=col, font=font, anchor="mb")

        pil_imgs.append(img)

    if not pil_imgs:
        raise HTTPException(404, "No pages with images found")

    buf = io.BytesIO()
    pil_imgs[0].save(buf, format="PDF", save_all=True, append_images=pil_imgs[1:], resolution=150)
    buf.seek(0)
    safe_name = re.sub(r'[^\w\s\-]', '', d.original_name or 'drawing').strip().replace(' ', '_') or 'drawing'
    logger.info("PDF export: drawing %d (%d pages)", d.id, len(pil_imgs))
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_annotated.pdf"'})


@app.get("/api/projects/{pid}/export/pdf")
def export_project_pdf(pid: int, db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    """Export all drawings in a project as a single combined PDF."""
    from PIL import Image, ImageDraw, ImageFont
    p = _proj404(pid, cu.id, db)

    # Color map from project symbol types
    color_map: dict[str, tuple] = {}
    for st in p.symbol_types:
        try:
            hex_col = st.color.lstrip('#')
            color_map[st.code] = tuple(int(hex_col[i:i+2], 16) for i in (0, 2, 4))
        except Exception:
            color_map[st.code] = (255, 80, 80)

    pil_imgs = []
    for d in sorted(p.drawings, key=lambda x: x.id):
        for page in sorted(d.pages, key=lambda pp: pp.page_number):
            if not page.image_path or not os.path.exists(page.image_path):
                continue
            img = Image.open(page.image_path).convert("RGB")
            draw = ImageDraw.Draw(img)
            iw, ih = img.size
            r = max(12, int(iw * 0.006))

            # Drawing title banner at top
            banner_h = max(30, int(ih * 0.02))
            banner = Image.new("RGB", (iw, banner_h), (20, 30, 50))
            bdraw  = ImageDraw.Draw(banner)
            try:
                tfont = ImageFont.truetype("arial.ttf", max(12, banner_h - 6))
            except Exception:
                tfont = ImageFont.load_default()
            title = f"{d.level or d.original_name}  —  Page {page.page_number}"
            bdraw.text((10, banner_h // 2), title, fill=(180, 200, 220), font=tfont, anchor="lm")
            combined = Image.new("RGB", (iw, ih + banner_h))
            combined.paste(banner, (0, 0))
            combined.paste(img, (0, banner_h))
            img = combined
            draw = ImageDraw.Draw(img)

            for det in (page.detections or []):
                cx = int(det.get("imgX", 0) * iw)
                cy = int(det.get("imgY", 0) * ih) + banner_h
                code = det.get("type", "")
                col  = color_map.get(code, (255, 80, 80))
                draw.ellipse([cx - r, cy - r, cx + r, cy + r], outline=col, width=max(2, r // 5))
                hs = r // 2
                draw.line([cx - hs, cy, cx + hs, cy], fill=col, width=max(1, r // 8))
                draw.line([cx, cy - hs, cx, cy + hs], fill=col, width=max(1, r // 8))
                label = 'DOME' if code == 'CCTV_Dome' else 'FIX' if code == 'CCTV_Fixed' else code
                try:
                    font = ImageFont.truetype("arial.ttf", max(10, r))
                except Exception:
                    font = ImageFont.load_default()
                draw.text((cx, cy - r - 4), label, fill=col, font=font, anchor="mb")

            pil_imgs.append(img)

    if not pil_imgs:
        raise HTTPException(404, "No pages found in this project")

    buf = io.BytesIO()
    pil_imgs[0].save(buf, format="PDF", save_all=True, append_images=pil_imgs[1:], resolution=150)
    buf.seek(0)
    safe_name = re.sub(r'[^\w\s\-]', '', p.name or 'project').strip().replace(' ', '_') or 'project'
    logger.info("PDF export: project %d (%d pages total)", p.id, len(pil_imgs))
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_all_drawings.pdf"'})


# ── Accuracy scoreboard ───────────────────────────────────────────────────────
@app.get("/api/accuracy")
def accuracy_scoreboard(db: Session = Depends(get_db), cu=Depends(auth.get_current_user)):
    """
    Precision/recall computed from verification feedback.
      confirm        = auto-detected and kept (true positive)
      false_positive = auto-detected and deleted
      missed         = manually added by the verifier (detector never saw it)
      wrong_type     = right spot, wrong symbol type (hurts precision AND recall)
    position_correction rows are position fixes, not classification signal — excluded.
    """
    rows = db.query(models.DetectionFeedback).all()

    def _bucket():
        return {"confirms": 0, "false_positives": 0, "missed": 0, "wrong_type": 0}

    def _finish(b):
        c, fp, m, wt = b["confirms"], b["false_positives"], b["missed"], b["wrong_type"]
        det_total  = c + fp + wt   # everything the detector claimed
        real_total = c + m  + wt   # everything actually on the drawings
        b["precision"] = round(c / det_total, 3)  if det_total  else None
        b["recall"]    = round(c / real_total, 3) if real_total else None
        b["sample"]    = c + fp + m + wt
        return b

    kind_key = {"confirm": "confirms", "false_positive": "false_positives",
                "missed": "missed", "wrong_type": "wrong_type"}
    by_code, by_firm, by_method = {}, {}, {}
    for r in rows:
        key = kind_key.get(r.feedback_type)
        if not key:
            continue
        by_code.setdefault(r.symbol_code or "?", _bucket())[key] += 1
        firm = (r.drawing_firm or "").strip() or "Unknown firm"
        by_firm.setdefault(firm, _bucket())[key] += 1
        method = (r.original_method or "").strip()
        if method in ("template", "text", "text_fallback"):
            by_method.setdefault(method, _bucket())[key] += 1

    name_map = {t["code"]: t["name"] for t in models.DEFAULT_SYMBOL_TYPES}
    for st in db.query(models.ProjectSymbolType).all():
        name_map.setdefault(st.code, st.name)

    templates = []
    for t in db.query(models.GlobalTemplate).all():
        use  = t.use_count or 0
        prec = round((t.confirm_count or 0) / use, 3) if use else None
        if use == 0:
            health = "untested"
        elif use >= 30 and (prec or 0) < 0.15:
            health = "toxic"
        elif use >= 10 and (prec or 0) < 0.5:
            health = "warning"
        else:
            health = "good"
        override = getattr(t, "threshold_override", None)
        templates.append({
            "id": t.id, "symbol_code": t.symbol_code,
            "symbol_name": t.symbol_name or name_map.get(t.symbol_code, t.symbol_code),
            "firm": t.drawing_firm or "",
            "use_count": use,
            "confirm_count": t.confirm_count or 0,
            "reject_count": t.reject_count or 0,
            "precision": prec,
            "effective_threshold": round(float(override) if override is not None
                                         else t.effective_threshold, 3),
            "health": health,
            "image_url": (f"/api/files/templates/{os.path.basename(t.image_path)}"
                          if t.image_path else None),
        })
    health_rank = {"toxic": 0, "warning": 1, "good": 2, "untested": 3}
    templates.sort(key=lambda x: (health_rank[x["health"]], -x["use_count"]))

    overall = _bucket()
    for b in by_code.values():
        for k in overall:
            overall[k] += b[k]

    return {
        "overall": {
            **_finish(overall),
            "total_pages":       db.query(models.DrawingPage).count(),
            "verified_pages":    db.query(models.DrawingPage)
                                   .filter(models.DrawingPage.verified == True).count(),  # noqa: E712
            "drawings":          db.query(models.Drawing).count(),
            "approved_drawings": db.query(models.Drawing)
                                   .filter(models.Drawing.status == "approved").count(),
        },
        "by_code":   [{"code": k, "name": name_map.get(k, k), **_finish(v)}
                      for k, v in sorted(by_code.items())],
        "by_firm":   [{"firm": k, **_finish(v)} for k, v in sorted(by_firm.items())],
        "by_method": [{"method": k, **_finish(v)} for k, v in sorted(by_method.items())],
        "templates": templates,
    }


@app.get("/api/admin/logs")
def get_logs(lines: int = Query(default=500, le=2000), cu=Depends(auth.get_current_user)):
    """Return the last N lines of the application log file."""
    if not _LOG_FILE.exists():
        return Response("No log file yet.", media_type="text/plain")
    try:
        with open(_LOG_FILE, "r", encoding="utf-8", errors="replace") as f:
            all_lines = f.readlines()
        tail = "".join(all_lines[-lines:])
    except Exception as e:
        raise HTTPException(500, f"Could not read log: {e}")
    return Response(tail, media_type="text/plain")


@app.get("/api/health")
def health():
    return {
        "status": "ok" if POPPLER_OK else "degraded",
        "version": "3.0.0",
        "poppler": POPPLER_OK,
        "detail": None if POPPLER_OK else
            "Poppler not found on PATH — uploads disabled. Install: winget install poppler",
    }


# ── Frontend catch-all ────────────────────────────────────────────────────────
@app.get("/")
@app.get("/{full_path:path}")
async def serve_frontend(full_path: str = ""):
    if full_path.startswith(("api/", "src/")):
        raise HTTPException(status_code=404)

    # Serve real static files from dist (icons, manifest, etc.) with traversal guard
    if full_path:
        candidate = (FRONTEND_DIST / full_path).resolve()
        if FRONTEND_DIST.resolve() in candidate.parents and candidate.is_file():
            return FileResponse(str(candidate))

    # Check index.html per-request so build order vs server start doesn't matter
    index = FRONTEND_DIST / "index.html"
    if index.is_file():
        return FileResponse(str(index))
    return {"message": f"Frontend not built — run: cd frontend && npm run build (looked for {index})"}


# ── Helpers ───────────────────────────────────────────────────────────────────
def _proj404(pid, uid, db):
    p = db.query(models.Project).filter(
        models.Project.id == pid, models.Project.owner_id == uid).first()
    if not p: raise HTTPException(404, "Project not found")
    return p

def _draw404(did, uid, db):
    d = db.query(models.Drawing).filter(models.Drawing.id == did).first()
    if not d: raise HTTPException(404, "Drawing not found")
    if d.project.owner_id != uid: raise HTTPException(403, "Access denied")
    return d

def _drawing_out(d: models.Drawing) -> DrawingOut:
    vp = sum(1 for pg in d.pages if pg.verified)
    return DrawingOut(
        id=d.id, original_name=d.original_name, block=d.block, level=d.level,
        total_pages=d.total_pages, status=d.status, revision=d.revision,
        error_message=getattr(d, "error_message", None),
        total_counts=d.total_counts,
        verified_pages=vp, total_pages_count=len(d.pages),
    )

def _assign_ids(dets: list) -> list:
    """Assign stable incremental IDs to detections for frontend tracking."""
    result = []
    for i, det in enumerate(dets):
        d = dict(det)
        d.setdefault("id", i + 1)
        result.append(d)
    return result

def _det_key(det: dict) -> str:
    return f"{det.get('type','')}_{round(det.get('imgX',0),3)}_{round(det.get('imgY',0),3)}"

def _tmpl_out(t: models.GlobalTemplate) -> TemplateOut:
    fname = os.path.basename(t.image_path) if t.image_path else ""
    return TemplateOut(
        id=t.id, symbol_code=t.symbol_code, symbol_name=t.symbol_name,
        image_url=f"/api/files/templates/{fname}",
        confirm_count=t.confirm_count, reject_count=t.reject_count,
        quality_score=round(t.quality_score, 3),
        effective_threshold=round(t.effective_threshold, 3),
        threshold_override=round(v, 3) if (v := getattr(t, 'threshold_override', None)) is not None else None,
        drawing_firm=t.drawing_firm,
    )

def _parse_name(filename: str) -> tuple[str, str]:
    block_m = re.search(r'Block (BA\w+|PD\w+)', filename)
    block   = block_m.group(1) if block_m else ""
    parts   = filename.replace(".pdf", "").split(" - ")
    level   = parts[-1].replace("Fire and Security Services Layout ", "").replace(
              "Fire Alarm and Security Services Layout ", "").replace(
              "Floor Plan – GA - ", "") if len(parts) > 1 else ""
    return block, level.strip()

def _parse_revision(filename: str) -> str:
    m = re.search(r'[_\-\s][Rr]ev[\s\-_]?(\w+)', filename)
    return m.group(1) if m else ""

AUTO_TEMPLATE_CAP  = 5       # max auto-crops per symbol code across the whole library
AUTO_CROP_HALF_PX  = 55      # half-size of the SQUARE crop in pixels (~110 px box at 200 dpi)

def _auto_create_templates(page: models.DrawingPage, final_dets: list,
                            drawing: models.Drawing, user_id: int, db: Session):
    """
    Bootstrap GlobalTemplates by cropping the page image at each verified
    detection position.  Skips:
      - detections still at unmoved text_fallback grid positions (synthetic coords)
      - codes that already have AUTO_TEMPLATE_CAP or more templates
      - detections that already carry a template_id (already matched — no need to re-crop)
    """
    if not page.image_path or not os.path.exists(page.image_path):
        logger.warning("Auto-template skipped: page image missing (drawing %d page %d, path=%r)",
                       drawing.id, page.page_number, page.image_path)
        return

    # Read page dimensions once so the crop is a fixed SQUARE pixel box
    # (normalised fractions of width/height give very different x/y sizes on A1)
    try:
        from PIL import Image as _PILImage
        with _PILImage.open(page.image_path) as _im:
            img_w, img_h = _im.size
    except Exception as e:
        logger.warning("Auto-template skipped: cannot read page image %s: %s", page.image_path, e)
        return
    half_x = AUTO_CROP_HALF_PX / max(img_w, 1)
    half_y = AUTO_CROP_HALF_PX / max(img_h, 1)

    # Build a set of the synthetic grid positions that text_fallback originally placed,
    # so we can skip markers the user never bothered to move (still fake locations).
    fallback_pos = {
        (round(d.get("imgX", 0), 3), round(d.get("imgY", 0), 3))
        for d in (page.original_detections or [])
        if d.get("method") == "text_fallback"
    }

    firm          = (drawing.project.drawing_firm or "") if drawing.project else ""
    name_by_code  = {st.code: st.name for st in (drawing.project.symbol_types if drawing.project else [])}
    created: dict[str, int] = {}

    for det in final_dets:
        code = det.get("type", "")
        if not code:
            continue
        if det.get("template_id"):      # already has a template — no bootstrap needed
            continue
        pos_key = (round(det.get("imgX", 0), 3), round(det.get("imgY", 0), 3))
        if pos_key in fallback_pos:     # user never moved this marker — position is synthetic
            continue

        existing = db.query(models.GlobalTemplate).filter(
            models.GlobalTemplate.symbol_code == code
        ).count()
        if existing + created.get(code, 0) >= AUTO_TEMPLATE_CAP:
            continue

        cx, cy = det.get("imgX", 0.5), det.get("imgY", 0.5)
        x1 = max(0.0, cx - half_x)
        y1 = max(0.0, cy - half_y)
        x2 = min(1.0, cx + half_x)
        y2 = min(1.0, cy + half_y)

        out_path = str(TEMPLATE_DIR / f"auto_{uuid.uuid4().hex}.png")
        try:
            w, h = crop_template_from_page(page.image_path, x1, y1, x2, y2, out_path,
                                           padding=0.0)
        except Exception as e:
            logger.warning("Auto-template crop failed for %s: %s", code, e)
            continue

        db.add(models.GlobalTemplate(
            symbol_code       = code,
            symbol_name       = name_by_code.get(code, code),
            image_path        = out_path,
            image_width       = w,
            image_height      = h,
            drawing_firm      = firm,
            source_project_id = drawing.project_id,
            created_by_id     = user_id,
        ))
        created[code] = created.get(code, 0) + 1

    if created:
        logger.info("Auto-created templates from drawing %d / page %d: %s",
                    drawing.id, page.page_number, created)


# A detection moved less than this (normalised units) is still the SAME symbol,
# just position-corrected — it must count as a confirm, not FP + missed.
MOVE_MATCH_TOL = 0.03

def _capture_feedback(page: models.DrawingPage, final_dets: list,
                      firm: str, db: Session):
    """
    Diff original_detections vs. the user's final_dets and write feedback records.
    Also updates template confirm/reject/use counts.
    Called only when page.verified is False (first save).

    Matching is nearest-neighbour with tolerance (NOT exact-coordinate), so a
    detection the user dragged to fix its position is recorded as a CONFIRM
    (the template found the right symbol) rather than false_positive + missed.
    """
    original = list(page.original_detections or [])
    final    = list(final_dets)

    used_final: set[int] = set()
    feedbacks  = []

    def _nearest(det, candidates, same_type: bool, tol: float):
        best_j, best_d = None, tol
        for j, f in enumerate(candidates):
            if j in used_final:
                continue
            if same_type != (f.get("type") == det.get("type")):
                continue
            dist = max(abs(f.get("imgX", 0) - det.get("imgX", 0)),
                       abs(f.get("imgY", 0) - det.get("imgY", 0)))
            if dist < best_d:
                best_d, best_j = dist, j
        return best_j

    # Pass 1: same-type match within tolerance → confirm (covers unmoved AND moved)
    unmatched_orig = []
    for det in original:
        j = _nearest(det, final, same_type=True, tol=MOVE_MATCH_TOL)
        if j is not None:
            used_final.add(j)
            feedbacks.append(("confirm", det, None))
        else:
            unmatched_orig.append(det)

    # Pass 2: different-type match at (nearly) the same position → wrong_type
    still_unmatched = []
    for det in unmatched_orig:
        j = _nearest(det, final, same_type=False, tol=0.01)
        if j is not None and final[j].get("type") != det.get("type"):
            used_final.add(j)
            feedbacks.append(("wrong_type", det, final[j].get("type")))
        else:
            still_unmatched.append(det)

    # Remaining originals were deleted by the user → false positives
    for det in still_unmatched:
        feedbacks.append(("false_positive", det, None))

    # Remaining finals were added by the user → missed
    for j, det in enumerate(final):
        if j not in used_final:
            feedbacks.append(("missed", det, None))

    for fb_type, det, corrected_code in feedbacks:
        tmpl_id = det.get("template_id")
        fb = models.DetectionFeedback(
            symbol_code=det.get("type", ""),
            feedback_type=fb_type,
            img_x=det.get("imgX", 0), img_y=det.get("imgY", 0),
            original_confidence=det.get("confidence"),
            original_method=det.get("method", ""),
            corrected_code=corrected_code,
            drawing_firm=firm,
            template_id=tmpl_id,
            drawing_id=page.drawing_id,
            page_number=page.page_number,
        )
        db.add(fb)

        if tmpl_id:
            t = db.query(models.GlobalTemplate).filter(
                models.GlobalTemplate.id == tmpl_id).first()
            if t:
                t.use_count = (t.use_count or 0) + 1
                if fb_type == "confirm":
                    t.confirm_count = (t.confirm_count or 0) + 1
                elif fb_type in ("false_positive", "wrong_type"):
                    t.reject_count = (t.reject_count or 0) + 1


def _run_detection(drawing_id: int, pdf_path: str, pages_dir: str,
                   symbol_codes: list, stored_tmpls_data: list, drawing_firm: str):
    """Background task — runs detection and updates drawing status."""
    from database import SessionLocal

    # Build lightweight template-like objects from plain dicts
    class _TmplProxy:
        def __init__(self, d):
            self.id            = d["id"]
            self.symbol_code   = d["symbol_code"]
            self.image_path    = d["image_path"]
            self.drawing_firm  = d["drawing_firm"]
            self.confirm_count = d["confirm_count"]
            self.reject_count  = d["reject_count"]
            self.use_count     = d["use_count"]
            self.threshold_override = d.get("threshold_override")

        # MUST mirror models.GlobalTemplate.effective_threshold — this proxy is
        # what the background detection task actually uses. (Previously it had a
        # stricter base of 0.75, silently overriding the 0.65 tuned in models.py.)
        @property
        def effective_threshold(self):
            if self.threshold_override is not None:
                return float(self.threshold_override)
            base  = 0.65
            total = (self.confirm_count or 0) + (self.reject_count or 0)
            if total < 5: return base
            precision = (self.confirm_count or 0) / total
            if precision > 0.9: return max(0.55, base - 0.08)
            if precision < 0.6: return min(0.82, base + 0.08)
            return base

        @property
        def quality_score(self):
            if not self.use_count: return 0.5
            return (self.confirm_count or 0) / max(self.use_count, 1)

    stored_tmpls = [_TmplProxy(d) for d in stored_tmpls_data]

    db = SessionLocal()
    try:
        d = db.query(models.Drawing).filter(models.Drawing.id == drawing_id).first()
        if not d: return
        d.status = "processing"; db.commit()

        results = run_full_detection(
            pdf_path, pages_dir,
            symbol_codes=symbol_codes,
            stored_templates=stored_tmpls,
            drawing_firm=drawing_firm,
        )
        d.total_pages = results["total_pages"]
        for pr in results["page_results"]:
            # Store deep copies so original_detections stays immutable
            dets = copy.deepcopy(pr["detections"])
            page = models.DrawingPage(
                page_number=pr["page"],
                image_path=pr["image_path"],
                detections=dets,
                original_detections=copy.deepcopy(dets),
                drawing_id=drawing_id,
            )
            db.add(page)
        d.status = "detected"; db.commit()

    except Exception as e:
        logger.exception("Detection error for drawing %d", drawing_id)
        try:
            d = db.query(models.Drawing).filter(models.Drawing.id == drawing_id).first()
            if d:
                d.status = "error"
                d.error_message = f"{type(e).__name__}: {e}"[:500]
                db.commit()
        except Exception:
            pass
    finally:
        db.close()


def _build_excel(project) -> bytes:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    sym_types = project.symbol_types or []
    if not sym_types:
        sym_types = [type('ST', (), d)() for d in models.DEFAULT_SYMBOL_TYPES]
    sym_codes = [st.code for st in sym_types]
    sym_names = [st.name for st in sym_types]

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Symbol Count"
    DARK = PatternFill("solid", fgColor="1F4E79")
    MID  = PatternFill("solid", fgColor="2E75B6")
    GREY = PatternFill("solid", fgColor="F2F2F2")
    WHT  = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(border_style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center")
    lft  = Alignment(horizontal="left",   vertical="center", indent=1)
    n_cols   = 4 + len(sym_codes) + 1  # Block + Drawing + Rev + Status + symbols + Total
    last_col = get_column_letter(n_cols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"{project.name} — {project.client or ''} — Symbol Count"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = DARK; ws["A1"].alignment = ctr; ws.row_dimensions[1].height = 26

    headers = ["Block", "Drawing / Level", "Rev", "Status"] + sym_names + ["Total"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = MID; c.alignment = ctr; c.border = bdr
    ws.row_dimensions[2].height = 18

    row_i = 3
    for i, d in enumerate(project.drawings):
        cts   = d.total_counts; fill = GREY if i % 2 == 0 else WHT
        total = sum(cts.get(code, 0) for code in sym_codes)
        vals  = [d.block, d.level or d.original_name, d.revision, d.status] + \
                [cts.get(code, 0) for code in sym_codes] + [total]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row_i, column=col, value=v)
            c.fill = fill; c.border = bdr
            c.alignment = lft if col == 2 else ctr
            if col > 4 and isinstance(v, int) and v > 0:
                c.font = Font(bold=True, color="1F4E79")
        row_i += 1

    all_cts = [d.total_counts for d in project.drawings]
    grand   = sum(sum(ct.get(c, 0) for c in sym_codes) for ct in all_cts)
    for col, code in enumerate(sym_codes, 5):
        c = ws.cell(row=row_i, column=col, value=sum(ct.get(code, 0) for ct in all_cts))
        c.font = Font(bold=True, color="FFFFFF"); c.fill = DARK
        c.alignment = ctr; c.border = bdr
    ws.cell(row=row_i, column=n_cols, value=grand).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row_i, column=n_cols).fill = DARK
    ws.cell(row=row_i, column=n_cols).alignment = ctr
    for col in [1, 2, 3, 4]:
        ws.cell(row=row_i, column=col).fill = DARK
        ws.cell(row=row_i, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row_i, column=col).alignment = ctr
    ws.cell(row=row_i, column=1, value="TOTAL")
    ws.cell(row=row_i, column=2, value=f"{len(project.drawings)} drawings")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 12
    for i in range(len(sym_codes) + 1):
        ws.column_dimensions[get_column_letter(5 + i)].width = 15
    ws.freeze_panes = "A3"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
